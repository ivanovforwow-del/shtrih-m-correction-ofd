# -*- coding: utf-8 -*-
"""
Парсер онлайн чеков с Платформы ОФД - версия 4
Исправлено: парсинг каждого товара отдельным блоком для корректного извлечения суммы
"""
import re
import csv
import time
import html as html_module
import sys
import requests
from datetime import datetime
from openpyxl import load_workbook

# Настройки
INPUT_FILE = 'Список чеков без НДС Финал.xlsx'
OUTPUT_FILE = 'receipts_data.csv'
SHEET_NAME = 'Лист2'

# Задержка между запросами (секунды)
REQUEST_DELAY = 0.3


def log(msg):
    """Логирование с немедленным выводом"""
    print(msg, flush=True)


def extract_receipt_params(hyperlink):
    """Извлекает параметры чека из гиперссылки Платформы ОФД"""
    if not hyperlink:
        return None
    
    params = {}
    id_match = re.search(r'id=(\d+)', hyperlink)
    if id_match:
        params['id'] = id_match.group(1)
    
    date_match = re.search(r'date=(\d+)', hyperlink)
    if date_match:
        params['date'] = date_match.group(1)
    
    fp_match = re.search(r'fp=(\d+)', hyperlink)
    if fp_match:
        params['fp'] = fp_match.group(1)
    
    return params if params else None


def fetch_receipt_data(params):
    """Получает данные чека с Платформы ОФД"""
    if not params or 'id' not in params:
        return None
    
    url = "https://lk.platformaofd.ru/web/noauth/cheque/id"
    
    try:
        response = requests.get(url, params={
            'id': params.get('id'),
            'date': params.get('date'),
            'fp': params.get('fp')
        }, timeout=15)
        
        if response.status_code == 200:
            return response.text
        else:
            return None
            
    except Exception as e:
        log(f"  Error fetching receipt: {e}")
        return None


def parse_receipt_items(html_content, fp):
    """
    Парсит HTML страницы чека и извлекает данные о товарах
    
    Возвращает список товаров с полями:
    - fp: фискальный признак
    - name: название товара
    - quantity: количество
    - unit: единица измерения (л, шт, кг)
    - price: цена за единицу
    - summ: сумма позиции
    - date: дата чека
    """
    items = []
    
    if not html_content:
        return items
    
    # Ищем блок fido_cheque_container
    match = re.search(r'id="fido_cheque_container">(.*?)</div>\s*<div', html_content, re.DOTALL)
    if not match:
        match = re.search(r'id="fido_cheque_container">(.*?)</div>$', html_content, re.DOTALL)
    
    if not match:
        return items
    
    # Декодируем HTML entities
    decoded_html = html_module.unescape(match.group(1))
    
    # Извлекаем дату из чека
    date_str = ""
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2})', decoded_html)
    if date_match:
        try:
            dt = datetime.strptime(date_match.group(1), '%d.%m.%Y %H:%M')
            date_str = dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            pass
    
    # Разбиваем HTML на блоки товаров
    # Каждый товар начинается с <b>НОМЕР: НАЗВАНИЕ</b>
    # И заканчивается перед следующим товаром или перед "ИТОГ"
    
    # Находим все блоки товаров
    # Паттерн: от <b>1: до следующего <b>2: или до ИТОГ
    item_blocks = re.split(r'<span>\s*<table[^>]*>.*?<b>(\d+):', decoded_html)
    
    # Альтернативный подход - находим каждый товар по номеру
    # Сначала найдем все номера товаров
    item_names = re.findall(r'<b>(\d+):\s*([^<]+)</b>', decoded_html)
    
    if not item_names:
        return items
    
    # Для каждого товара найдем его блок в HTML
    for i, (num, name_raw) in enumerate(item_names):
        item = {
            'fp': fp,
            'name': name_raw.strip(),
            'quantity': 0,
            'price': 0,
            'summ': 0,
            'unit': 'шт',
            'date': date_str
        }
        
        # Находим блок текущего товара
        # Ищем от "<b>NUM:" до следующего "<b>NUM+1:" или "ИТОГ"
        current_pattern = rf'<b>{num}:\s*[^<]+</b>'
        next_num = int(num) + 1
        next_pattern = rf'<b>{next_num}:\s*|ИТОГ|<b>ИТОГ'
        
        current_match = re.search(current_pattern, decoded_html)
        if not current_match:
            items.append(item)
            continue
        
        start_pos = current_match.start()
        
        # Ищем конец блока
        rest_html = decoded_html[start_pos:]
        next_match = re.search(rf'<b>{next_num}:\s*|ИТОГ', rest_html[50:])  # Пропускаем текущий заголовок
        
        if next_match:
            block_html = rest_html[:50 + next_match.start()]
        else:
            block_html = rest_html[:2000]  # Ограничиваем длину блока
        
        # Извлекаем данные из блока
        
        # 1. Сумма - ищем после "Общая стоимость позиции с учетом скидок и наценок"
        summ_match = re.search(
            r'Общая стоимость позиции[^<]*</span>\s*</td>\s*<td[^>]*>\s*<span[^>]*>([\d.]+)</span>',
            block_html
        )
        if summ_match:
            try:
                item['summ'] = float(summ_match.group(1))
            except:
                pass
        
        # 2. Цена - ищем паттерн "QTY UNIT x PRICE"
        price_match = re.search(
            r'<span>(\d+\.?\d*)\s*</span>\s*<span>\s*<span>(л|шт\.?|кг)</span>\s*</span>\s*x\s*<span>([\d.]+)</span>',
            block_html
        )
        if price_match:
            try:
                item['quantity'] = float(price_match.group(1))
                item['unit'] = 'шт' if price_match.group(2) in ['шт', 'шт.'] else price_match.group(2)
                item['price'] = float(price_match.group(3))
            except:
                pass
        
        # 3. Проверяем паттерн топлива в названии: "АИ-95-К5 (45 л * 66.70)"
        fuel_pattern = r'^(.+?)\s*\((\d+\.?\d*)\s*(л|шт|кг)?\s*\*\s*(\d+\.?\d*)\)$'
        fuel_match = re.match(fuel_pattern, item['name'])
        
        if fuel_match:
            item['name'] = fuel_match.group(1).strip()
            item['quantity'] = float(fuel_match.group(2))
            item['unit'] = fuel_match.group(3) if fuel_match.group(3) else 'л'
            item['price'] = float(fuel_match.group(4))
        else:
            # Проверяем паттерн с количеством в названии: "Товар (5 шт * 129.00)"
            qty_in_name = re.match(r'^(.+?)\s*\((\d+\.?\d*)\s*(л|шт|кг)?\s*\*\s*(\d+\.?\d*)\)$', name_raw)
            if qty_in_name:
                item['name'] = qty_in_name.group(1).strip()
                item['quantity'] = float(qty_in_name.group(2))
                item['unit'] = qty_in_name.group(3) if qty_in_name.group(3) else 'шт'
                item['price'] = float(qty_in_name.group(4))
        
        items.append(item)
    
    return items


def read_excel_data():
    """Читает данные из Excel файла"""
    log(f"Reading Excel file: {INPUT_FILE}")
    
    wb = load_workbook(INPUT_FILE, data_only=False)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    
    headers = [cell.value for cell in ws[1]]
    
    # Находим индекс колонки с гиперссылкой
    hyperlink_col = None
    for idx, h in enumerate(headers):
        if h and ('чек' in str(h).lower() or 'посмотреть' in str(h).lower()):
            hyperlink_col = idx
            break
    
    if hyperlink_col is None:
        hyperlink_col = len(headers) - 1  # Последняя колонка
    
    receipts = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        cell = row[hyperlink_col]
        hyperlink = None
        
        if cell.hyperlink:
            hyperlink = cell.hyperlink.target
        elif cell.data_type == 'f' and cell.value:
            hyperlink = str(cell.value)
        
        if hyperlink:
            receipts.append({
                'row': row_idx,
                'hyperlink': hyperlink
            })
    
    wb.close()
    log(f"Total rows with hyperlinks: {len(receipts)}")
    
    return receipts


def main():
    """Главная функция"""
    log("=" * 60)
    log("ПАРСЕР ОНЛАЙН ЧЕКОВ С ПЛАТФОРМЫ ОФД (версия 4)")
    log("=" * 60)
    
    # Читаем Excel
    receipts = read_excel_data()
    
    # Собираем данные о товарах
    all_items = []
    processed = 0
    errors = 0
    skipped = 0
    
    total = len(receipts)
    log(f"\nProcessing {total} receipts...")
    log("")
    
    for i, receipt in enumerate(receipts):
        # Выводим прогресс каждые 50 чеков
        if i % 50 == 0:
            progress = (i / total) * 100
            log(f"Progress: {i}/{total} ({progress:.1f}%) - Items: {len(all_items)}")
        
        # Извлекаем параметры
        params = extract_receipt_params(receipt['hyperlink'])
        
        if params:
            fp = params.get('fp')
            
            # Получаем данные чека
            html_content = fetch_receipt_data(params)
            
            if html_content:
                # Парсим товары
                items = parse_receipt_items(html_content, fp)
                
                if items:
                    for item in items:
                        item['row'] = receipt['row']
                        all_items.append(item)
                    processed += 1
                else:
                    errors += 1
                    log(f"  No items found for FP={fp}")
                
                # Задержка между запросами
                time.sleep(REQUEST_DELAY)
            else:
                errors += 1
        else:
            skipped += 1
    
    # Сохраняем в CSV
    log("")
    log("=" * 60)
    log(f"Total items found: {len(all_items)}")
    log(f"Receipts processed: {processed}")
    log(f"Errors: {errors}")
    log(f"Skipped: {skipped}")
    
    if all_items:
        with open(OUTPUT_FILE, 'w', newline='', encoding='utf-8-sig') as f:
            fieldnames = ['row', 'fp', 'name', 'quantity', 'unit', 'price', 'summ', 'date']
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            for item in all_items:
                writer.writerow(item)
        
        log(f"Saved to: {OUTPUT_FILE}")
        
        # Показываем примеры
        log("\nFirst 10 items:")
        for item in all_items[:10]:
            log(f"  {item['name']}: {item['quantity']} {item['unit']} x {item['price']} = {item['summ']}")
        
        # Проверка корректности данных
        log("\n=== ПРОВЕРКА КОРРЕКТНОСТИ ===")
        errors_count = 0
        for item in all_items:
            if item['quantity'] > 0 and item['price'] > 0:
                calculated = round(item['quantity'] * item['price'], 2)
                if abs(calculated - item['summ']) > 0.02:
                    errors_count += 1
                    if errors_count <= 10:
                        log(f"  ERROR: {item['name']}: {item['quantity']} x {item['price']} = {calculated} != {item['summ']}")
        
        if errors_count == 0:
            log("  All items passed validation!")
        else:
            log(f"  Total errors: {errors_count}")
    else:
        log("No items found!")


if __name__ == "__main__":
    main()
