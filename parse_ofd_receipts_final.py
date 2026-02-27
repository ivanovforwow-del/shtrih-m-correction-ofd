# -*- coding: utf-8 -*-
"""
Финальный парсер онлайн чеков с Платформы ОФД
Корректно извлекает данные о товарах: количество, единицу измерения, цену и сумму
"""
import re
import csv
import time
import html as html_module
import sys
import requests
from datetime import datetime
from openpyxl import load_workbook

# Настройки
INPUT_FILE = 'Список чеков без НДС Финал.xlsx'
OUTPUT_FILE = 'receipts_data.csv'
SHEET_NAME = 'Лист2'

# Задержка между запросами (секунды)
REQUEST_DELAY = 0.3


def log(msg):
    """Логирование с немедленным выводом"""
    print(msg, flush=True)


def extract_receipt_params(hyperlink):
    """Извлекает параметры чека из гиперссылки Платформы ОФД"""
    if not hyperlink:
        return None
    
    params = {}
    id_match = re.search(r'id=(\d+)', hyperlink)
    if id_match:
        params['id'] = id_match.group(1)
    
    date_match = re.search(r'date=(\d+)', hyperlink)
    if date_match:
        params['date'] = date_match.group(1)
    
    fp_match = re.search(r'fp=(\d+)', hyperlink)
    if fp_match:
        params['fp'] = fp_match.group(1)
    
    return params if params else None


def fetch_receipt_data(params):
    """Получает данные чека с Платформы ОФД"""
    if not params or 'id' not in params:
        return None
    
    url = "https://lk.platformaofd.ru/web/noauth/cheque/id"
    
    try:
        response = requests.get(url, params={
            'id': params.get('id'),
            'date': params.get('date'),
            'fp': params.get('fp')
        }, timeout=15)
        
        if response.status_code == 200:
            return response.text
        else:
            return None
            
    except Exception as e:
        log(f"  Error fetching receipt: {e}")
        return None


def parse_receipt_items(html_content, fp):
    """
    Парсит HTML страницы чека и извлекает данные о товарах
    
    Возвращает список товаров с полями:
    - fp: фискальный признак
    - name: название товара
    - quantity: количество
    - unit: единица измерения (л, шт, кг)
    - price: цена за единицу
    - summ: сумма позиции
    - date: дата чека
    - marking_code: код маркировки товара (например, EAN-13)
    - marking_type: тип маркировки (MarkingType)
    - marking_type2: тип КМ (MarkingType2)
    """
    items = []
    
    if not html_content:
        return items
    
    # Ищем блок fido_cheque_container
    match = re.search(r'id="fido_cheque_container">(.*?)</div>\s*<div', html_content, re.DOTALL)
    if not match:
        match = re.search(r'id="fido_cheque_container">(.*?)</div>$', html_content, re.DOTALL)
    
    if not match:
        return items
    
    # Декодируем HTML entities
    decoded_html = html_module.unescape(match.group(1))
    
    # Извлекаем дату из чека
    date_str = ""
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2})', decoded_html)
    if date_match:
        try:
            dt = datetime.strptime(date_match.group(1), '%d.%m.%Y %H:%M')
            date_str = dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            pass
    
    # Находим все названия товаров: <b>1: НАЗВАНИЕ</b>
    item_names = re.findall(r'<b>(\d+):\s*([^<]+)</b>', decoded_html)
    
    # Находим все строки с ценой: <span>1 </span> <span><span>шт.</span></span> x <span>220.00</span>
    price_lines = re.findall(
        r'<span>(\d+\.?\d*)\s*</span>\s*<span>\s*<span>(л|шт\.?|кг)</span>\s*</span>\s*x\s*<span>([\d.]+)</span>',
        decoded_html
    )
    
    # Находим все суммы "Общая стоимость позиции"
    # Сумма находится в <span> после "Общая стоимость позиции"
    parts = re.split(r'Общая стоимость позиции', decoded_html)
    sums = []
    for part in parts[1:]:  # Пропускаем первую часть до первого маркера
        summ_match = re.search(r'<span[^>]*>([\d.]+)</span>', part)
        if summ_match:
            sums.append(summ_match.group(1))
    
    # Находим коды маркировки для каждого товара
    # Коды маркировки находятся в блоке товара, например: <span>КТ EAN-13</span> <span>8801056791018</span>
    # Паттерн: после "EAN-13" идет значение в следующем span
    marking_codes = []
    
    # Ищем коды маркировки по паттерну: EAN-13</span></td><td...><span>код</span>
    # Это позволяет отличить маркировку от других 13-значных номеров (номер ФН, ККТ и т.д.)
    all_codes = re.findall(r'EAN-13</span>\s*</td>\s*<td[^>]*>\s*<span[^>]*>(\d{13})</span>', decoded_html)
    
    # Создаем список кодов по порядку товаров
    for i in range(len(item_names)):
        if i < len(all_codes):
            marking_codes.append(all_codes[i])
        else:
            marking_codes.append("")
    
    # Функция для определения MarkingType по коду маркировки
    def get_marking_type(code):
        if not code:
            return 0
        length = len(code)
        if length == 8:
            return 17672  # EAN-8
        elif length == 13:
            return 17677  # EAN-13
        elif length == 14:
            return 18702  # ITF-14
        else:
            return 0  # Нераспознанный вид кода
    
    # Функция для определения MarkingType2 (тип КМ)
    def get_marking_type2(code):
        if not code:
            return 0  # Код маркировки отсутствует
        # Для простоты пока используем только базовые значения
        # В реальности нужно проверять формат кода и длину контрольной суммы
        return 0
    
    # Объединяем данные
    for i, (num, name_raw) in enumerate(item_names):
        item = {
            'fp': fp,
            'name': name_raw.strip(),
            'quantity': 0,
            'price': 0,
            'summ': 0,
            'unit': 'шт',
            'date': date_str,
            'marking_code': marking_codes[i] if i < len(marking_codes) else "",
            'marking_type': get_marking_type(marking_codes[i] if i < len(marking_codes) else ""),
            'marking_type2': get_marking_type2(marking_codes[i] if i < len(marking_codes) else "")
        }
        
        # Сумма
        if i < len(sums):
            try:
                item['summ'] = float(sums[i])
            except:
                pass
        
        # Паттерн 1: Топливо - данные в названии
        # "АИ-95-К5 (45 л * 66.70)"
        fuel_pattern = r'^(.+?)\s*\((\d+\.?\d*)\s*(л|шт|кг)?\s*\*\s*(\d+\.?\d*)\)$'
        fuel_match = re.match(fuel_pattern, item['name'])
        
        if fuel_match:
            item['name'] = fuel_match.group(1).strip()
            item['quantity'] = float(fuel_match.group(2))
            item['unit'] = fuel_match.group(3) if fuel_match.group(3) else 'л'
            item['price'] = float(fuel_match.group(4))
        else:
            # Паттерн 2: Штучный товар - берем цену из price_lines
            if i < len(price_lines):
                qty, unit, price = price_lines[i]
                item['quantity'] = float(qty)
                item['unit'] = 'шт' if unit in ['шт', 'шт.'] else unit
                item['price'] = float(price)
                
                # Проверяем: если quantity=1 и price != summ
                # Это случай когда показана цена за единицу, а не общая
                if item['quantity'] == 1 and abs(item['price'] - item['summ']) > 0.01:
                    # Ищем количество в названии
                    qty_in_name = re.search(r'\((\d+\.?\d*)\s*(л|шт)', name_raw)
                    if qty_in_name:
                        item['quantity'] = float(qty_in_name.group(1))
                        item['unit'] = qty_in_name.group(2)
                        item['price'] = round(item['summ'] / item['quantity'], 2)
        
        items.append(item)
    
    return items


def read_excel_data():
    """Читает данные из Excel файла"""
    log(f"Reading Excel file: {INPUT_FILE}")
    
    wb = load_workbook(INPUT_FILE, data_only=False)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    
    headers = [cell.value for cell in ws[1]]
    
    # Находим индекс колонки с гиперссылкой
    hyperlink_col = None
    for idx, h in enumerate(headers):
        if h and ('чек' in str(h).lower() or 'посмотреть' in str(h).lower()):
            hyperlink_col = idx
            break
    
    if hyperlink_col is None:
        hyperlink_col = len(headers) - 1  # Последняя колонка
    
    receipts = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        cell = row[hyperlink_col]
        hyperlink = None
        
        if cell.hyperlink:
            hyperlink = cell.hyperlink.target
        elif cell.data_type == 'f' and cell.value:
            hyperlink = str(cell.value)
        
        if hyperlink:
            receipts.append({
                'row': row_idx,
                'hyperlink': hyperlink
            })
    
    wb.close()
    log(f"Total rows with hyperlinks: {len(receipts)}")
    
    return receipts


def main():
    """Главная функция"""
    log("=" * 60)
    log("ПАРСЕР ОНЛАЙН ЧЕКОВ С ПЛАТФОРМЫ ОФД (финальная версия)")
    log("=" * 60)
    
    # Читаем Excel
    receipts = read_excel_data()
    
    # Собираем данные о товарах
    all_items = []
    processed = 0
    errors = 0
    skipped = 0
    
    total = len(receipts)
    log(f"\nProcessing {total} receipts...")
    log("")
    
    for i, receipt in enumerate(receipts):
        # Выводим прогресс каждые 50 чеков
        if i % 50 == 0:
            progress = (i / total) * 100
            log(f"Progress: {i}/{total} ({progress:.1f}%) - Items: {len(all_items)}")
        
        # Извлекаем параметры
        params = extract_receipt_params(receipt['hyperlink'])
        
        if params:
            fp = params.get('fp')
            
            # Получаем данные чека
            html_content = fetch_receipt_data(params)
            
            if html_content:
                # Парсим товары
                items = parse_receipt_items(html_content, fp)
                
                if items:
                    for item in items:
                        item['row'] = receipt['row']
                        all_items.append(item)
                    processed += 1
                else:
                    errors += 1
                    log(f"  No items found for FP={fp}")
                
                # Задержка между запросами
                time.sleep(REQUEST_DELAY)
            else:
                errors += 1
        else:
            skipped += 1
    
    # Сохраняем в CSV
    log("")
    log("=" * 60)
    log(f"Total items found: {len(all_items)}")
    log(f"Receipts processed: {processed}")
    log(f"Errors: {errors}")
    log(f"Skipped: {skipped}")
    
    if all_items:
        with open(OUTPUT_FILE, 'w', newline='', encoding='utf-8-sig') as f:
            fieldnames = ['row', 'fp', 'name', 'quantity', 'unit', 'price', 'summ', 'date', 'marking_code', 'marking_type', 'marking_type2']
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            for item in all_items:
                writer.writerow(item)
        
        log(f"Saved to: {OUTPUT_FILE}")
        
        # Показываем примеры
        log("\nFirst 10 items:")
        for item in all_items[:10]:
            log(f"  {item['name']}: {item['quantity']} {item['unit']} x {item['price']} = {item['summ']}")
        
        # Проверка корректности данных
        log("\n=== ПРОВЕРКА КОРРЕКТНОСТИ ===")
        errors_count = 0
        for item in all_items:
            if item['quantity'] > 0 and item['price'] > 0:
                calculated = round(item['quantity'] * item['price'], 2)
                if abs(calculated - item['summ']) > 0.02:
                    errors_count += 1
                    if errors_count <= 10:
                        log(f"  ERROR: {item['name']}: {item['quantity']} x {item['price']} = {calculated} != {item['summ']}")
        
        if errors_count == 0:
            log("  All items passed validation!")
        else:
            log(f"  Total errors: {errors_count}")
    else:
        log("No items found!")


if __name__ == "__main__":
    main()
