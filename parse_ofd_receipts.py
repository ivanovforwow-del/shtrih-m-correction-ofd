# -*- coding: utf-8 -*-
"""
Парсер онлайн чеков с Платформы ОФД
Извлекает данные о товарах из чеков по ссылке
"""
import re
import csv
import time
import html
import sys
import requests
from datetime import datetime
from openpyxl import load_workbook

# Настройки
INPUT_FILE = 'Список чеков без НДС Финал.xlsx'
OUTPUT_FILE = 'receipts_data.csv'
SHEET_NAME = 'Лист2'

# Задержка между запросами
REQUEST_DELAY = 0.3


def log(msg):
    """Логирование с немедленным выводом"""
    print(msg, flush=True)


def extract_receipt_params(hyperlink):
    """Извлекает параметры чека из гиперссылки Платформы ОФД"""
    if not hyperlink:
        return None
    
    params = {}
    id_match = re.search(r'id=(\d+)', hyperlink)
    if id_match:
        params['id'] = id_match.group(1)
    
    date_match = re.search(r'date=(\d+)', hyperlink)
    if date_match:
        params['date'] = date_match.group(1)
    
    fp_match = re.search(r'fp=(\d+)', hyperlink)
    if fp_match:
        params['fp'] = fp_match.group(1)
    
    return params if params else None


def fetch_receipt_data(params):
    """Получает данные чека с Платформы ОФД"""
    if not params or 'id' not in params:
        return None
    
    url = "https://lk.platformaofd.ru/web/noauth/cheque/id"
    
    try:
        response = requests.get(url, params={
            'id': params.get('id'),
            'date': params.get('date'),
            'fp': params.get('fp')
        }, timeout=10)
        
        if response.status_code == 200:
            return response.text
        else:
            return None
            
    except Exception as e:
        return None


def parse_receipt_items(html_content, fp):
    """
    Парсит HTML страницы чека и извлекает данные о товарах
    """
    items = []
    
    if not html_content:
        return items
    
    # Ищем блок fido_cheque_container
    match = re.search(r'id="fido_cheque_container">(.*?)</div>', html_content, re.DOTALL)
    if not match:
        return items
    
    # Декодируем HTML entities
    decoded_html = html.unescape(match.group(1))
    
    # Ищем товары
    # Паттерн: <b>1: ДТ-А-К5 (0.004 л * 92.50)</b>
    name_pattern = r'<b>(\d+):\s*([^<]+)</b>'
    name_matches = re.findall(name_pattern, decoded_html)
    
    for num, name in name_matches:
        item = {
            'fp': fp,
            'name': name.strip(),
            'quantity': 0,
            'price': 0,
            'summ': 0,
            'unit': 'л'
        }
        
        # Парсим название товара с количеством и ценой
        # Формат: "ДТ-А-К5 (0.004 л * 92.50)"
        name_parts = re.match(r'([^(]+)\s*\((\d+\.?\d*)\s*(л|шт|кг)?\s*\*\s*(\d+\.?\d*)\)', name)
        if name_parts:
            item['name'] = name_parts.group(1).strip()
            item['quantity'] = float(name_parts.group(2))
            if name_parts.group(3):
                item['unit'] = name_parts.group(3)
            item['price'] = float(name_parts.group(4))
            item['summ'] = round(item['quantity'] * item['price'], 2)
        
        items.append(item)
    
    return items


def read_excel_data():
    """Читает данные из Excel файла"""
    log(f"Reading Excel file: {INPUT_FILE}")
    
    wb = load_workbook(INPUT_FILE, data_only=False)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    
    headers = [cell.value for cell in ws[1]]
    
    receipts = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        row_data = {}
        for col_idx, cell in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
            row_data[header] = cell.value
            
            if cell.hyperlink:
                row_data['hyperlink'] = cell.hyperlink.target
            elif cell.data_type == 'f' and 'HYPERLINK' in str(cell.value).upper():
                row_data['hyperlink_formula'] = cell.value
        
        receipts.append({
            'row': row_idx,
            'data': row_data
        })
    
    wb.close()
    log(f"Total rows: {len(receipts)}")
    
    return receipts


def main():
    """Главная функция"""
    log("=" * 60)
    log("ПАРСЕР ОНЛАЙН ЧЕКОВ С ПЛАТФОРМЫ ОФД")
    log("=" * 60)
    
    # Читаем Excel
    receipts = read_excel_data()
    
    # Собираем данные о товарах
    all_items = []
    processed = 0
    errors = 0
    
    total = len(receipts)
    log(f"\nProcessing {total} receipts...")
    log("")
    
    for i, receipt in enumerate(receipts):
        # Выводим прогресс каждые 10 чеков
        if i % 10 == 0:
            progress = (i / total) * 100
            log(f"Progress: {i}/{total} ({progress:.1f}%) - Items: {len(all_items)}")
        
        # Извлекаем гиперссылку
        hyperlink = receipt['data'].get('hyperlink_formula') or receipt['data'].get('hyperlink')
        
        if hyperlink:
            # Извлекаем параметры
            params = extract_receipt_params(hyperlink)
            
            if params:
                fp = params.get('fp')
                
                # Получаем данные чека
                html_content = fetch_receipt_data(params)
                
                if html_content:
                    # Парсим товары
                    items = parse_receipt_items(html_content, fp)
                    
                    if items:
                        for item in items:
                            item['row'] = receipt['row']
                            all_items.append(item)
                        processed += 1
                    else:
                        errors += 1
                    
                    # Задержка между запросами
                    time.sleep(REQUEST_DELAY)
    
    # Сохраняем в CSV
    log("")
    log("=" * 60)
    log(f"Total items found: {len(all_items)}")
    log(f"Receipts processed: {processed}")
    log(f"Errors: {errors}")
    
    if all_items:
        with open(OUTPUT_FILE, 'w', newline='', encoding='utf-8-sig') as f:
            fieldnames = ['row', 'fp', 'name', 'quantity', 'unit', 'price', 'summ']
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            for item in all_items:
                writer.writerow(item)
        
        log(f"Saved to: {OUTPUT_FILE}")
        
        # Показываем примеры
        log("\nFirst 10 items:")
        for item in all_items[:10]:
            log(f"  {item['name']}: {item['quantity']} {item['unit']} x {item['price']} = {item['summ']}")
    else:
        log("No items found!")


if __name__ == "__main__":
    main()
