# -*- coding: utf-8 -*-
"""
Скачивание тестовых чеков для анализа
"""
import requests
import re
from openpyxl import load_workbook

INPUT_FILE = 'Список чеков без НДС Финал.xlsx'
SHEET_NAME = 'Лист2'

# Читаем xlsx
wb = load_workbook(INPUT_FILE, data_only=False)
ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]

# Ищем нужные чеки
target_fps = ['1852047038', '1175267103']  # Чеки с несколькими товарами

for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
    cell = row[8]  # Колонка с гиперссылкой
    if cell.data_type == 'f':
        formula = str(cell.value)
        fp_match = re.search(r'fp=(\d+)', formula)
        id_match = re.search(r'id=(\d+)', formula)
        date_match = re.search(r'date=(\d+)', formula)
        
        if fp_match and id_match:
            fp = fp_match.group(1)
            if fp in target_fps:
                receipt_id = id_match.group(1)
                date = date_match.group(1)
                
                print(f'Row {row_idx}: FP={fp}')
                
                url = 'https://lk.platformaofd.ru/web/noauth/cheque/id'
                response = requests.get(url, params={'id': receipt_id, 'date': date, 'fp': fp}, timeout=10)
                
                with open(f'response_{fp}.html', 'w', encoding='utf-8') as f:
                    f.write(response.text)
                print(f'  Saved: response_{fp}.html ({len(response.text)} bytes)')
                
                target_fps.remove(fp)
                if not target_fps:
                    break

wb.close()
print('Done!')
