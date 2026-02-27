# -*- coding: utf-8 -*-
"""
Улучшенный парсер онлайн чеков с Платформы ОФД
Корректно извлекает данные о товарах: количество, единицу измерения, цену и сумму
"""
import re
import csv
import time
import html as html_module
import sys
import requests
from datetime import datetime
from openpyxl import load_workbook

# Настройки
INPUT_FILE = 'Список чеков без НДС Финал.xlsx'
OUTPUT_FILE = 'receipts_data.csv'
SHEET_NAME = 'Лист2'

# Задержка между запросами (секунды)
REQUEST_DELAY = 0.3


def log(msg):
    """Логирование с немедленным выводом"""
    print(msg, flush=True)


def extract_receipt_params(hyperlink):
    """Извлекает параметры чека из гиперссылки Платформы ОФД"""
    if not hyperlink:
        return None
    
    params = {}
    id_match = re.search(r'id=(\d+)', hyperlink)
    if id_match:
        params['id'] = id_match.group(1)
    
    date_match = re.search(r'date=(\d+)', hyperlink)
    if date_match:
        params['date'] = date_match.group(1)
    
    fp_match = re.search(r'fp=(\d+)', hyperlink)
    if fp_match:
        params['fp'] = fp_match.group(1)
    
    return params if params else None


def fetch_receipt_data(params):
    """Получает данные чека с Платформы ОФД"""
    if not params or 'id' not in params:
        return None
    
    url = "https://lk.platformaofd.ru/web/noauth/cheque/id"
    
    try:
        response = requests.get(url, params={
            'id': params.get('id'),
            'date': params.get('date'),
            'fp': params.get('fp')
        }, timeout=15)
        
        if response.status_code == 200:
            return response.text
        else:
            return None
            
    except Exception as e:
        log(f"  Error fetching receipt: {e}")
        return None


def parse_receipt_items(html_content, fp):
    """
    Парсит HTML страницы чека и извлекает данные о товарах
    
    Структура HTML:
    1. Топливо: <b>1: АИ-95-К5 (45 л * 66.70)</b>
       - Название, количество, единица, цена в одной строке
       - Сумма: "Общая стоимость позиции... 3001.50"
    
    2. Штучный товар: <b>2: Капучино (Время Кофе) с корицей 300мл</b>
       - Только название
       - Количество и цена: "1 шт. x 220.00"
       - Сумма: "Общая стоимость позиции... 220.00"
    """
    items = []
    
    if not html_content:
        return items
    
    # Ищем блок fido_cheque_container
    match = re.search(r'id="fido_cheque_container">(.*?)</div>\s*<div', html_content, re.DOTALL)
    if not match:
        # Пробуем альтернативный паттерн
        match = re.search(r'id="fido_cheque_container">(.*?)</div>$', html_content, re.DOTALL)
    
    if not match:
        return items
    
    # Декодируем HTML entities
    decoded_html = html_module.unescape(match.group(1))
    
    # Извлекаем дату из чека
    date_str = ""
    date_match = re.search(r'(\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2})', decoded_html)
    if date_match:
        try:
            dt = datetime.strptime(date_match.group(1), '%d.%m.%Y %H:%M')
            date_str = dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            pass
    
    # Разбиваем на блоки товаров (между <!-- Предоплата --> и <!-- Итого -->)
    # Каждый товар в блоке <span>...</span>
    
    # Паттерн для поиска блоков товаров
    # Ищем: <span> <table ...> ... <b>1: НАЗВАНИЕ</b> ... </table> </span>
    item_blocks = re.findall(
        r'<span>\s*<table[^>]*>.*?<b>(\d+):\s*([^<]+)</b>.*?'
        r'Общая стоимость позиции.*?<span>([\d.]+)</span>.*?</table>\s*</span>',
        decoded_html, 
        re.DOTALL
    )
    
    for num, name_raw, summ in item_blocks:
        item = {
            'fp': fp,
            'name': name_raw.strip(),
            'quantity': 0,
            'price': 0,
            'summ': float(summ),
            'unit': 'шт',
            'date': date_str
        }
        
        # Паттерн 1: Топливо - данные в названии
        # "АИ-95-К5 (45 л * 66.70)"
        fuel_pattern = r'^(.+?)\s*\((\d+\.?\d*)\s*(л|шт|кг)?\s*\*\s*(\d+\.?\d*)\)$'
        fuel_match = re.match(fuel_pattern, item['name'])
        
        if fuel_match:
            item['name'] = fuel_match.group(1).strip()
            item['quantity'] = float(fuel_match.group(2))
            item['unit'] = fuel_match.group(3) if fuel_match.group(3) else 'л'
            item['price'] = float(fuel_match.group(4))
        else:
            # Паттерн 2: Штучный товар - ищем цену в блоке
            # Ищем: "1 шт. x 220.00" или "1 шт x 220.00" или "1 л x 3001.50"
            # В блоке товара ищем строку с ценой
            
            # Находим полный блок товара для извлечения цены
            block_pattern = rf'<b>{num}:\s*{re.escape(name_raw)}</b>.*?</table>\s*</span>'
            block_match = re.search(block_pattern, decoded_html, re.DOTALL)
            
            if block_match:
                block_html = block_match.group(0)
                
                # Ищем цену: <span>1 </span> <span> <span>шт.</span> </span> x <span>220.00</span>
                # Или: <span>1 </span> <span> <span>л</span> </span> x <span>3001.50</span>
                price_pattern = r'<span>(\d+\.?\d*)\s*</span>\s*<span>\s*<span>(л|шт\.?|кг)</span>\s*</span>\s*x\s*<span>([\d.]+)</span>'
                price_match = re.search(price_pattern, block_html)
                
                if price_match:
                    item['quantity'] = float(price_match.group(1))
                    unit_raw = price_match.group(2).strip('.')
                    item['unit'] = 'шт' if unit_raw in ['шт', 'шт.'] else unit_raw
                    item['price'] = float(price_match.group(3))
                else:
                    # Альтернативный паттерн для топлива (когда цена за литр)
                    # <span>1 </span> <span> <span>л</span> </span> x <span>3001.50</span>
                    alt_pattern = r'<span>(\d+\.?\d*)\s*</span>\s*<span>\s*<span>(л|шт\.?|кг)</span>\s*</span>\s*x\s*<span>([\d.]+)</span>'
                    alt_match = re.search(alt_pattern, block_html)
                    if alt_match:
                        item['quantity'] = float(alt_match.group(1))
                        item['unit'] = alt_match.group(2).strip('.')
                        item['price'] = float(alt_match.group(3))
        
        # Проверяем корректность: quantity * price должно быть равно summ
        if item['quantity'] > 0 and item['price'] > 0:
            calculated = round(item['quantity'] * item['price'], 2)
            # Если рассчитанная сумма не совпадает, возможно цена за единицу
            if abs(calculated - item['summ']) > 0.01:
                # Это случай когда показана цена за единицу, а не общая
                # Например: 1 л x 3001.50 - это 1 литр по цене 3001.50 за литр
                # Но на самом деле 45 литров по 66.70 = 3001.50
                # В этом случае нужно пересчитать
                if item['quantity'] == 1:
                    # Ищем реальное количество в названии
                    qty_in_name = re.search(r'\((\d+\.?\d*)\s*(л|шт)', name_raw)
                    if qty_in_name:
                        item['quantity'] = float(qty_in_name.group(1))
                        item['unit'] = qty_in_name.group(2)
                        item['price'] = round(item['summ'] / item['quantity'], 2)
        
        items.append(item)
    
    return items


def read_excel_data():
    """Читает данные из Excel файла"""
    log(f"Reading Excel file: {INPUT_FILE}")
    
    wb = load_workbook(INPUT_FILE, data_only=False)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
    
    headers = [cell.value for cell in ws[1]]
    
    # Находим индекс колонки с гиперссылкой
    hyperlink_col = None
    for idx, h in enumerate(headers):
        if h and ('чек' in str(h).lower() or 'посмотреть' in str(h).lower()):
            hyperlink_col = idx
            break
    
    if hyperlink_col is None:
        hyperlink_col = len(headers) - 1  # Последняя колонка
    
    receipts = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        cell = row[hyperlink_col]
        hyperlink = None
        
        if cell.hyperlink:
            hyperlink = cell.hyperlink.target
        elif cell.data_type == 'f' and cell.value:
            hyperlink = str(cell.value)
        
        if hyperlink:
            receipts.append({
                'row': row_idx,
                'hyperlink': hyperlink
            })
    
    wb.close()
    log(f"Total rows with hyperlinks: {len(receipts)}")
    
    return receipts


def main():
    """Главная функция"""
    log("=" * 60)
    log("УЛУЧШЕННЫЙ ПАРСЕР ОНЛАЙН ЧЕКОВ С ПЛАТФОРМЫ ОФД")
    log("=" * 60)
    
    # Читаем Excel
    receipts = read_excel_data()
    
    # Собираем данные о товарах
    all_items = []
    processed = 0
    errors = 0
    skipped = 0
    
    total = len(receipts)
    log(f"\nProcessing {total} receipts...")
    log("")
    
    for i, receipt in enumerate(receipts):
        # Выводим прогресс каждые 50 чеков
        if i % 50 == 0:
            progress = (i / total) * 100
            log(f"Progress: {i}/{total} ({progress:.1f}%) - Items: {len(all_items)}")
        
        # Извлекаем параметры
        params = extract_receipt_params(receipt['hyperlink'])
        
        if params:
            fp = params.get('fp')
            
            # Получаем данные чека
            html_content = fetch_receipt_data(params)
            
            if html_content:
                # Парсим товары
                items = parse_receipt_items(html_content, fp)
                
                if items:
                    for item in items:
                        item['row'] = receipt['row']
                        all_items.append(item)
                    processed += 1
                else:
                    errors += 1
                    log(f"  No items found for FP={fp}")
                
                # Задержка между запросами
                time.sleep(REQUEST_DELAY)
            else:
                errors += 1
        else:
            skipped += 1
    
    # Сохраняем в CSV
    log("")
    log("=" * 60)
    log(f"Total items found: {len(all_items)}")
    log(f"Receipts processed: {processed}")
    log(f"Errors: {errors}")
    log(f"Skipped: {skipped}")
    
    if all_items:
        with open(OUTPUT_FILE, 'w', newline='', encoding='utf-8-sig') as f:
            fieldnames = ['row', 'fp', 'name', 'quantity', 'unit', 'price', 'summ', 'date']
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            for item in all_items:
                writer.writerow(item)
        
        log(f"Saved to: {OUTPUT_FILE}")
        
        # Показываем примеры
        log("\nFirst 10 items:")
        for item in all_items[:10]:
            log(f"  {item['name']}: {item['quantity']} {item['unit']} x {item['price']} = {item['summ']}")
        
        # Проверка корректности данных
        log("\n=== ПРОВЕРКА КОРРЕКТНОСТИ ===")
        errors_count = 0
        for item in all_items:
            if item['quantity'] > 0 and item['price'] > 0:
                calculated = round(item['quantity'] * item['price'], 2)
                if abs(calculated - item['summ']) > 0.02:
                    errors_count += 1
                    if errors_count <= 10:
                        log(f"  ERROR: {item['name']}: {item['quantity']} x {item['price']} = {calculated} != {item['summ']}")
        
        if errors_count == 0:
            log("  All items passed validation!")
        else:
            log(f"  Total errors: {errors_count}")
    else:
        log("No items found!")


if __name__ == "__main__":
    main()
