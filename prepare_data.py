import pandas as pd
import re
import csv
from pathlib import Path
from openpyxl import load_workbook

# Настройки
INPUT_FILE = 'Список чеков без НДС Финал.xlsx'
OUTPUT_FILE = 'list.csv'
SHEET_NAME = 'Лист2'  # Название листа из вашего файла

# Названия колонок для фильтрации НДС 10%
# Возможные варианты: 'НДС', 'Ставка НДС', 'НДС %' и т.д.
VAT_COLUMNS = ['НДС', 'Ставка НДС', 'НДС %', 'Ставка']

def extract_fp_from_hyperlink(hyperlink_formula):
    """
    Извлекает фискальный признак (fp) из формулы ГИПЕРССЫЛКА.
    Пример: =ГИПЕРССЫЛКА("...&fp=2703755211";"...")
    """
    if not isinstance(hyperlink_formula, str) or 'fp=' not in hyperlink_formula:
        return None
    
    # Ищем паттерн fp=ЧИСЛО
    match = re.search(r'fp=(\d+)', hyperlink_formula)
    if match:
        return match.group(1)
    return None


def find_vat_column(df):
    """
    Находит колонку с НДС в DataFrame.
    Возвращает название колонки или None, если не найдена.
    """
    for col in df.columns:
        col_lower = str(col).lower().strip()
        # Проверяем точное название "НДС 10%" или похожие
        if 'ндс' in col_lower or 'налог' in col_lower or 'vat' in col_lower:
            return col
    return None


def is_vat_10(row, vat_column):
    """
    Проверяет, содержит ли строка НДС 10%.
    """
    if vat_column is None:
        return False
    
    vat_value = row.get(vat_column)
    if pd.isna(vat_value):
        return False
    
    # Проверяем числовое значение 0.1 (это 10%)
    try:
        vat_float = float(vat_value)
        # 0.1 = 10%, также возможны варианты 10, 10.0
        if abs(vat_float - 0.1) < 0.001 or abs(vat_float - 10) < 0.1:
            return True
    except (ValueError, TypeError):
        # Если это строка
        vat_str = str(vat_value).lower().strip()
        return '10' in vat_str or '10%' in vat_str or 'десять' in vat_str
    
    return False

def prepare_csv():
    print(f"Reading file: {INPUT_FILE}")
    
    # Читаем Excel файл
    try:
        df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=0)
    except Exception as e:
        print(f"Error reading Excel: {e}")
        # Пробуем прочитать первый лист, если указанный не найден
        df = pd.read_excel(INPUT_FILE, sheet_name=0, header=0)
        print("Using first sheet")
    
    print(f"Found rows: {len(df)}")
    
    # Выводим структуру колонок для отладки
    print(f"\nAvailable columns: {list(df.columns)}")
    
    # Ищем колонку с НДС 10%
    vat_column = None
    for col in df.columns:
        col_str = str(col).strip()
        if col_str == 'НДС 10%':
            vat_column = col_str
            break
    
    if vat_column:
        print(f"Found VAT 10% column: {vat_column}")
    else:
        print("VAT 10% column not found!")
    
    # Проверяем наличие нужных колонок
    required_columns = ['Дата/время', 'Наличными', 'Электронными', 'Посмотреть чек']
    for col in required_columns:
        if col not in df.columns:
            print(f"Error: Column '{col}' not found")
            print(f"Available columns: {list(df.columns)}")
            return
    
    # Отладочный вывод - первые 3 строки
    print("\n=== DEBUG: First 3 rows ===")
    for idx in range(min(3, len(df))):
        row = df.iloc[idx]
        print(f"\nRow {idx+2}:")
        print(f"  Дата/время: {row.get('Дата/время')}")
        print(f"  Наличными: {row.get('Наличными')} (type: {type(row.get('Наличными'))})")
        print(f"  Электронными: {row.get('Электронными')} (type: {type(row.get('Электронными'))})")
        print(f"  Итого: {row.get('Итого')} (type: {type(row.get('Итого'))})")
        print(f"  Посмотреть чек: {row.get('Посмотреть чек')}")
        
        # Также проверим Без НДС
        print(f"  Без НДС: {row.get('Без НДС')} (type: {type(row.get('Без НДС'))})")
    print("=== END DEBUG ===\n")
    
    # Читаем гиперссылки через openpyxl
    print("\nReading hyperlinks via openpyxl...")
    hyperlinks = {}
    try:
        # Читаем с формулами
        wb = load_workbook(INPUT_FILE, data_only=False, keep_vba=True)
        # Получаем лист по имени или по индексу
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.worksheets[0]
        
        # Находим индекс колонки "Посмотреть чек"
        hyperlink_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Посмотреть чек':
                hyperlink_col_idx = idx
                break
        
        if hyperlink_col_idx:
            print(f"Hyperlink column index: {hyperlink_col_idx}")
            # Читаем гиперссылки (начиная со 2-й строки)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                cell = row[hyperlink_col_idx - 1]
                # Проверяем гиперссылку
                if cell.hyperlink:
                    hyperlinks[row_idx] = cell.hyperlink.target
                # Проверяем формулу
                elif cell.data_type == 'f':
                    # Это формула
                    formula = cell.value
                    if formula and 'HYPERLINK' in str(formula).upper():
                        hyperlinks[row_idx] = str(formula)
                # Иначе проверяем значение ячейки
                elif cell.value and 'fp=' in str(cell.value):
                    hyperlinks[row_idx] = str(cell.value)
        
        wb.close()
        print(f"Found {len(hyperlinks)} hyperlinks")
        
    except Exception as e:
        print(f"Error reading hyperlinks: {e}")
    
    # Подготавливаем данные для CSV
    processed_data = []
    skipped_rows = 0
    skipped_vat_10 = 0
    
    for index, row in df.iterrows():
        excel_row = index + 2  # +2 because index starts from 0 and we have header
        
        # Проверяем НДС 10% - пропускаем такие чеки
        if vat_column:
            vat_value = row.get(vat_column)
            if pd.notna(vat_value):
                try:
                    vat_float = float(vat_value)
                    # Если есть значение > 0 в колонке НДС 10%, значит там был НДС 10%
                    if vat_float > 0:
                        skipped_vat_10 += 1
                        continue
                except (ValueError, TypeError):
                    pass
        
        # Определяем тип оплаты
        cash = float(row['Наличными']) if pd.notna(row['Наличными']) else 0
        electronic = float(row['Электронными']) if pd.notna(row['Электронными']) else 0
        
        # Определяем тип оплаты для нашей системы (0 - безнал, 1 - нал)
        if cash > 0 and electronic == 0:
            payment_type = 1  # наличные
            summ = cash
        elif electronic > 0 and cash == 0:
            payment_type = 0  # безналичные
            summ = electronic
        else:
            # Если в чеке смешанная оплата (и нал, и безнал)
            # ВАЖНО: В коррекции такие чеки нужно разбивать на два!
            print(f"  ! Row {excel_row}: mixed payment (cash:{cash}, card:{electronic}) - needs manual processing")
            skipped_rows += 1
            continue
        
        # Извлекаем фискальный признак из ссылки
        hyperlink = hyperlinks.get(excel_row, '')
        fp = extract_fp_from_hyperlink(hyperlink)
        
        if not fp:
            # Пробуем взять из самого Excel
            link_value = row.get('Посмотреть чек')
            if pd.notna(link_value):
                fp = extract_fp_from_hyperlink(str(link_value))
        
        if not fp:
            # print(f"  ! Row {excel_row}: cannot extract FP from link")
            skipped_rows += 1
            continue
        
        # Добавляем обработанную строку
        processed_data.append({
            'summ': f"{summ:.2f}",
            'type': payment_type,
            'fiscal_sign': fp
        })
    
    print(f"\nSuccessfully processed: {len(processed_data)} checks")
    print(f"Skipped (VAT 10%): {skipped_vat_10}")
    print(f"Skipped (mixed payment or errors): {skipped_rows}")
    
    # Сохраняем в CSV
    with open(OUTPUT_FILE, 'w', newline='', encoding='utf-8-sig') as csvfile:
        fieldnames = ['summ', 'type', 'fiscal_sign']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=';')
        
        writer.writeheader()
        for data in processed_data:
            writer.writerow(data)
    
    print(f"\nФайл сохранен: {OUTPUT_FILE}")
    print("\nПервые 5 строк для проверки:")
    print("summ;type;fiscal_sign")
    for i, data in enumerate(processed_data[:5]):
        print(f"{data['summ']};{data['type']};{data['fiscal_sign']}")

if __name__ == "__main__":
    # Настройка кодировки для Windows
    import sys
    if sys.platform == 'win32':
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    # Проверяем наличие pandas
    try:
        import pandas as pd
    except ImportError:
        print("Ошибка: Не установлена библиотека pandas.")
        print("Установите её командой: pip install pandas openpyxl")
        exit(1)
    
    prepare_csv()